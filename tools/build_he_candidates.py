# -*- coding: utf-8 -*-
"""Build the Hebrew answer bank from the Wikipedia frequency corpus.

There are ~74k well-formed 5-letter candidates and we want ~2,500, so every
filter here is tuned for PRECISION, not recall: it is fine to drop a good word,
because thousands more are waiting behind it.

Filters, in order:
  1. exactly 5 Hebrew letters, no final-form letter out of final position
  2. attested in the corpus (this alone kills fabricated words like חיהים)
  3. not a particle-prefixed form  (בשנות = ב+שנות)
  4. not a construct/smichut form  (מלחמת, חיילי)
  5. morphological evidence of being an ordinary word: it must take the
     definite article or a plural/feminine ending somewhere in the corpus.
     Proper nouns overwhelmingly don't (מדריד, ראובן, ורסאי score 0).
  6. an explicit stoplist for the handful of proper nouns that survive (5),
     e.g. ישראל, which passes only because "ישראלים" exists.
"""
import openpyxl, json, sys
sys.stdout.reconfigure(encoding="utf-8")

SRC = r"C:\Users\assaf\Downloads\רשימת-המילים-בשפה-העברית-ושכיחות-על-פי-ויקיפדיה.xlsx"
HEB = set("אבגדהוזחטיכלמנסעפצקרשתךםןףץ")
FINALS = set("ךםןףץ")
FINAL_TO_MED = {"ך": "כ", "ם": "מ", "ן": "נ", "ף": "פ", "ץ": "צ"}

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
F = {}
for word, freq in ws.iter_rows(values_only=True):
    if isinstance(word, str) and isinstance(freq, (int, float)):
        w = word.strip()
        if w and all(c in HEB for c in w):
            F[w] = int(freq)

def stem(w):
    return w[:-1] + FINAL_TO_MED.get(w[-1], w[-1])

PREFIXES = "בהוכלמש"
def particle_form(w):
    if w[0] in PREFIXES and F.get(w[1:], 0) >= 25:
        return True
    if w[0] in "וש" and F.get(w[1:], 0) >= 25:
        return True
    return False

def construct_form(w):
    # מלחמת <- מלחמה ,  חיילי <- חיילים
    if w.endswith("ת") and F.get(w[:-1] + "ה", 0) >= 400:
        return True
    if w.endswith("י") and F.get(w[:-1] + "ים", 0) >= 250:
        return True
    return False


def alt_spelling(w):
    # גרסא/נוסחא/קופסא: non-standard א for a word normally written with ה
    return w.endswith("א") and F.get(w[:-1] + "ה", 0) >= 100

def future_plural(w):
    # תקבלו / תפתחו / תכנסו — 2nd-person future, a poor answer
    return w.startswith("ת") and w.endswith("ו")

def evidence(w):
    s = stem(w)
    return sum(F.get(v, 0) for v in (
        "ה" + w, s + "ים", s + "ות", s + "ה", "ה" + s + "ים", "ה" + s + "ות",
    ))

# Frequent proper nouns that clear the morphology bar anyway.
STOPLIST = {
    "ישראל", "מיכאל", "אירופה", "אמריקה", "אנגליה", "צרפתי", "גרמניה",
    "יהודים", "יהודית", "ירושלים", "אנגלית", "רוסיה", "איטליה", "ספרדי",
    "בריטי", "יהודי", "ערבית", "אנגלי", "רוסית", "גרמני", "איטלקי",
    "יוונית", "צרפתית", "ישראלי", "יהודה", "שמעון", "יעקב", "אברהם",
    "מרים", "אסתר", "דניאל", "גבריאל", "רפאל", "אוריאל", "נתניה",
    "חיפה", "אשדוד", "רחובות", "הרצל", "בגין", "רבין", "פרס",
    # places
    "הולנד", "מצרים", "לבנון", "עיראק", "איראן", "קמרון", "פולין",
    "רוסיה", "סוריה", "ברזיל", "בלגיה", "כווית", "עומאן", "מרוקו",
    "אלסקה", "שיקגו", "ברלין", "מדריד", "אתונה", "טורקיה", "יוגוסלביה",
    "אשכנז", "הבלטי", "אוקראינה", "ארגנטינה", "וייטנאם", "אוסטרליה",
    "שבדיה", "נורבגיה", "דנמרק", "פורטוגל", "מקסיקו", "קולומביה",
    "ונציה", "מילאנו", "בוליביה", "טוסקנה", "בוואריה", "סהרה",
    # personal names
    "אלברט", "אוסקר", "פטריק", "קתרין", "ארתור", "אליהו", "נחמיה",
    "ראובן", "אנדרו", "רוברט", "תומאס", "אדוארד", "פיליפ", "סטיבן",
    "מרטין", "הרולד", "אלפרד", "סרגיי", "אנטון", "בוריס", "אולגה",
    "נטליה", "סופיה", "אלעזר", "עקיבא", "שמואל", "אהרון", "יהושע",
    "וויליאם", "ריצרד", "צארלס", "אנתוני", "ניקולס", "אלכסנדר",
    "פרידריך", "היינריך", "וילהלם", "יוהאן", "לואיס", "קרלוס",
    "פרננדו", "ריקרדו", "מריו", "לוקאס", "גבעתי", "גולני", "צאנז",
}

candidates = []
for w, c in F.items():
    if len(w) != 5 or any(ch in FINALS for ch in w[:4]):
        continue
    if w in STOPLIST:
        continue
    if particle_form(w) or construct_form(w):
        continue
    if alt_spelling(w) or future_plural(w):
        continue
    if evidence(w) < 100:
        continue
    candidates.append((w, c, evidence(w)))

candidates.sort(key=lambda t: -t[1])
print(f"candidates passing all filters: {len(candidates)}")

for floor in (1,5,10,20,30,50,80,120):
    n=sum(1 for _,c,_ in candidates if c>=floor)
    print(f"  freq >= {floor:>4}: {n} words")

FLOOR = 20
final = [t for t in candidates if t[1] >= FLOOR]
print(f"taking top {len(final)} by corpus frequency")
print(f"  most frequent:  {final[0][0]} ({final[0][1]})")
print(f"  least frequent: {final[-1][0]} ({final[-1][1]})")

json.dump([[w, c] for w, c, _ in final],
          open("bank_candidates.json", "w", encoding="utf-8"), ensure_ascii=False)

for lo in range(len(final)-250, len(final), 250):
    print(f"\n--- {lo}-{lo+250} ---")
    print(" ".join(w for w, _, _ in final[lo:lo+250]))
